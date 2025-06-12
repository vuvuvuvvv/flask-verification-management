from flask import Blueprint, jsonify, send_file, request
from flask_jwt_extended import jwt_required
from app.models import DongHo
from app import db
from werkzeug.exceptions import NotFound
import json
from app.utils.url_encrypt import decode
from app.utils.pdf_generator import process_excel_bytesio_to_pdf
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Border, Side, Alignment, Font
from app.constants import PP_THUC_HIEN
from datetime import datetime
import traceback

from io import BytesIO
import shutil

import os

download_bp = Blueprint("download", __name__)

class DuLieuMotLanChay:
    def __init__(self, V1: float, V2: float, Vc1: float, Vc2: float):
        self.V1 = V1
        self.V2 = V2
        self.Vc1 = Vc1
        self.Vc2 = Vc2

def _get_sai_so_dong_ho(form_value: DuLieuMotLanChay) -> float:
    if form_value:
        # Kiểm tra và chuyển đổi các giá trị sang float
        try:
            V1 = float(form_value.V1) if form_value.V1 not in [None, ''] else 0.0
            V2 = float(form_value.V2) if form_value.V2 not in [None, ''] else 0.0
            Vc1 = float(form_value.Vc1) if form_value.Vc1 not in [None, ''] else 0.0
            Vc2 = float(form_value.Vc2) if form_value.Vc2 not in [None, ''] else 0.0
        except ValueError:
            return None  # Hoặc có thể trả về một giá trị mặc định khác

        if V2 == 0 and Vc2 == 0 and Vc1 == 0:
            return None

        VDHCT = V2 - V1
        VDHC = Vc2 - Vc1
        if VDHC != 0:
            error = ((VDHCT - VDHC) / VDHC) * 100
            return round(error, 3)  # Làm tròn đến 3 chữ số thập phân

    return None

def _create_sample_bb_excel_buffer_file(dongho: DongHo, is_preview: bool = False):
    row_heights = 0.69 #cm
    try:
        dongho_dict = dongho.to_dict()
        if "du_lieu_kiem_dinh" in dongho_dict:
            try:
                dongho_dict["du_lieu_kiem_dinh"] = json.loads(
                    dongho_dict["du_lieu_kiem_dinh"]
                )
            except json.JSONDecodeError as e:
                print("Error decoding JSON: ", str(e))
                return None, None
        if dongho_dict['du_lieu_kiem_dinh']['ket_qua'] is None:
            print(">>> Không có dữ liệu kiểm định!")
            return None, None
        
        fileName = (
            "KĐ_BB"
            + ("_" + dongho.so_giay_chung_nhan if dongho.so_giay_chung_nhan else "")
            + ("_" + dongho.ten_khach_hang if dongho.ten_khach_hang else "")
            # + ("_" + dongho.ten_dong_ho if dongho.ten_dong_ho else "")
            + ("_DN" + dongho.dn if dongho.dn else "")
            + ("_" + dongho.ngay_thuc_hien.strftime("%d-%m-%Y") if dongho.ngay_thuc_hien else "")
            + ".xlsx"
        )

        bb_file_path = f"files/export/BB/{fileName}"
        bb_directory = os.path.dirname(bb_file_path)

        # Ensure the directory exists
        if not os.path.exists(bb_directory):
            os.makedirs(bb_directory)

        if not os.path.exists(bb_file_path):
            suffix = "Preview" if is_preview else "ExcelForm"
            src_file = f"files/BB_{suffix}.xlsm"
            workbook = openpyxl.load_workbook(src_file)
            sheet = workbook.active

            positions = ["22", "23", "24", "26"]

            for pos in positions:
                check_img = Image("files/data/check.png")
                sheet.add_image(check_img, f"U{pos}")
                check_img.width = 25
                check_img.height = 25

                circle_img = Image("files/data/circle.png")
                sheet.add_image(circle_img, f"Z{pos}")
                circle_img.width = 25
                circle_img.height = 25

            sheet.cell(row=2, column=20, value="BIÊN BẢN KIỂM ĐỊNH")

            if dongho.so_giay_chung_nhan:
                sheet.cell(
                    row=3, column=20, value=f"FMS.KĐ.{dongho.so_giay_chung_nhan}.{dongho.ngay_thuc_hien.strftime('%y') if dongho.ngay_thuc_hien else '?'}"
                )

            sheet.cell(
                row=5, column=8, value=dongho.ten_phuong_tien_do if dongho.ten_phuong_tien_do else ""
            )
            sheet.cell(
                row=6,
                column=8,
                value=dongho.co_so_san_xuat if dongho.co_so_san_xuat else "",
            )

            sheet.cell(
                row=7,
                column=8,
                value=(
                    (
                        f"Sensor: {dongho.sensor}"
                        if dongho.transitor
                        else dongho.sensor
                    )
                    if dongho.sensor
                    else ""
                ),
            )

            sheet.cell(
                row=8,
                column=8,
                value=(
                    f"Tranmistter: {dongho.transitor}" if dongho.transitor else ""
                ),
            )
            # TODO: seri_
            # sheet.cell(
            #     row=7,
            #     column=27,
            #     value=(
            #         (
            #             f"Sensor: {dongho.seri_sensor}"
            #             if dongho.seri_chi_thi
            #             else dongho.seri_sensor
            #         )
            #         if dongho.seri_sensor
            #         else ""
            #     ),
            # )

            # sheet.cell(
            #     row=8,
            #     column=27,
            #     value=(
            #         f"Tranmistter: {dongho.seri_chi_thi}" if dongho.seri_chi_thi else ""
            #     ),
            # )

            if dongho.dn:
                sheet.cell(row=9, column=26, value=dongho.dn)

            if dongho.q3 or dongho.qn:
                sheet.cell(row=10, column=24, value="Q3=" if dongho.q3 else "Qn=")
                sheet.cell(
                    row=10, column=25, value=dongho.q3 if dongho.q3 else dongho.qn
                )

            # M12
            if dongho.ccx:
                sheet.cell(row=11, column=13, value="-")
                sheet.cell(row=11, column=14, value=f"Cấp chính xác: {dongho.ccx}")

            if dongho.so_qd_pdm:
                sheet.cell(row=12, column=13, value="-")
                sheet.cell(row=12, column=14, value=f"Ký hiệu PDM / Số quyết định:")
                sheet.cell(row=12, column=26, value=dongho.so_qd_pdm)

            # R19
            if dongho.k_factor:
                sheet.cell(row=13, column=13, value="-")
                sheet.cell(row=13, column=14, value=f"Hệ số K:")
                sheet.cell(row=13, column=19, value=dongho.k_factor)
            
            # # H7
            if dongho.co_so_su_dung:
                sheet.cell(row=14, column=8, value=dongho.noi_su_dung)

            # K10
            if dongho.phuong_phap_thuc_hien:
                sheet.cell(row=16, column=11, value=dongho.phuong_phap_thuc_hien)

            if dongho.chuan_thiet_bi_su_dung:
                sheet.cell(row=17, column=11, value=dongho.chuan_thiet_bi_su_dung)

            if dongho.nguoi_thuc_hien:
                sheet.cell(row=19, column=9, value=str(dongho.nguoi_thuc_hien).upper())

            sheet.cell(
                row=19,
                column=29,
                value=(
                    dongho.ngay_thuc_hien.strftime("%d/%m/%Y")
                    if dongho.ngay_thuc_hien
                    else ""
                ),
            )

            sheet[f"R31"] = "°C"
            sheet[f"AE31"] = "°C"

            sheet.cell(
                row=20,
                column=2,
                value=f"Địa điểm thực hiện: {dongho.dia_diem_thuc_hien}" if dongho.dia_diem_thuc_hien else "",
            )

            desired_height = row_heights * 28.35
            # Run from row: 2
            for row in range(2, sheet.max_row + 1):
                sheet.row_dimensions[row].height = desired_height

            # Bảng dl: start from row 32:
            du_lieu = dongho_dict['du_lieu_kiem_dinh']['du_lieu']
            hieu_sai_so = list(dongho_dict['du_lieu_kiem_dinh']['hieu_sai_so'])
            titles = ["Q3","Q2","Q1"] if dongho.q3 else ['Qn', "Qt", "Qmin"] 

            black_solid = Side(style="thin", color="000000")   # Viền liền màu đen (mỏng)
            black_dotted = Side(style="dotted", color="000000")

            start_row = 32
            center_align = Alignment(horizontal="center", vertical="center")
            font_12 = Font(size=12, color="000000")
            font_13 = Font(size=13, color="000000")

            for index, ll in enumerate(titles):
                tmp_start_row = start_row
                ll_display = {"Q1": "I", "Qmin": "I", "Q2": "II", "Qt": "II", "Q3": "III", "Qn": "III"}.get(ll, ll)
                dl_ll = du_lieu[ll]
                lan_chay = dict(dl_ll['lan_chay'])

                sheet[f"B{start_row}"] = ll_display
                sheet[f"B{start_row}"].alignment = center_align
                sheet[f"B{start_row}"].font = font_12
                sheet[f"B{start_row}"].number_format = 'General'

                value = 0.3 * float(dl_ll['value']) if ll == "Q3" else float(dl_ll['value'])
                sheet[f"D{start_row}"] = round(value, 2) if value is not None else 0.0
                sheet[f"D{start_row}"].alignment = center_align
                sheet[f"D{start_row}"].font = font_12
                sheet[f"D{start_row}"].number_format = 'General'

                sheet.merge_cells(f"B{start_row}:C{start_row + len(lan_chay) - 1}")
                sheet.merge_cells(f"D{start_row}:F{start_row + len(lan_chay) - 1}")
                sheet.merge_cells(f"AJ{start_row}:AL{start_row + len(lan_chay) - 1}")

                hss = None

                for key, val in lan_chay.items():
                    def get_float(v): return float(v) if v not in [None, ''] else 0.0

                    sheet.merge_cells(f"G{start_row}:J{start_row}")
                    sheet[f"G{start_row}"] = get_float(val['V1'])
                    sheet[f"G{start_row}"].alignment = center_align
                    sheet[f"G{start_row}"].font = font_12
                    sheet[f"G{start_row}"].number_format = 'General'

                    sheet.merge_cells(f"K{start_row}:N{start_row}")
                    sheet[f"K{start_row}"] = get_float(val['V2'])
                    sheet[f"K{start_row}"].alignment = center_align
                    sheet[f"K{start_row}"].font = font_12
                    sheet[f"K{start_row}"].number_format = 'General'

                    sheet.merge_cells(f"O{start_row}:Q{start_row}")
                    try:
                        delta_v = round(get_float(val['V2']) - get_float(val['V1']), 2) if val not in [None, ''] and val['V2'] and val['V1'] and (get_float(val['V2']) - get_float(val['V1']) != 0) else 0.0
                        sheet[f"O{start_row}"] = delta_v
                    except ValueError:
                        sheet[f"O{start_row}"] = "-"
                    sheet[f"O{start_row}"].alignment = center_align
                    sheet[f"O{start_row}"].font = font_12
                    sheet[f"O{start_row}"].number_format = 'General'

                    sheet.merge_cells(f"R{start_row}:S{start_row}")
                    sheet[f"R{start_row}"] = get_float(val['Tdh'])
                    sheet[f"R{start_row}"].alignment = center_align
                    sheet[f"R{start_row}"].font = font_12
                    sheet[f"R{start_row}"].number_format = 'General'

                    sheet.merge_cells(f"T{start_row}:W{start_row}")
                    sheet[f"T{start_row}"] = get_float(val['Vc1'])
                    sheet[f"T{start_row}"].alignment = center_align
                    sheet[f"T{start_row}"].font = font_12
                    sheet[f"T{start_row}"].number_format = 'General'

                    sheet.merge_cells(f"X{start_row}:AA{start_row}")
                    sheet[f"X{start_row}"] = get_float(val['Vc2'])
                    sheet[f"X{start_row}"].alignment = center_align
                    sheet[f"X{start_row}"].font = font_12
                    sheet[f"X{start_row}"].number_format = 'General'

                    sheet.merge_cells(f"AB{start_row}:AD{start_row}")
                    try:
                        delta_vc = round(get_float(val['Vc2']) - get_float(val['Vc1']), 2) if val and val['Vc2'] and val['Vc1'] and get_float(val['Vc2']) - get_float(val['Vc1']) != 0 else 0.0
                        sheet[f"AB{start_row}"] = delta_vc
                    except ValueError:
                        sheet[f"AB{start_row}"] = "-"
                    sheet[f"AB{start_row}"].alignment = center_align
                    sheet[f"AB{start_row}"].font = font_12
                    sheet[f"AB{start_row}"].number_format = 'General'

                    sheet.merge_cells(f"AE{start_row}:AF{start_row}")
                    sheet[f"AE{start_row}"] = get_float(val['Tc'])
                    sheet[f"AE{start_row}"].alignment = center_align
                    sheet[f"AE{start_row}"].font = font_12
                    sheet[f"AE{start_row}"].number_format = 'General'

                    sheet.merge_cells(f"AG{start_row}:AI{start_row}")
                    try:
                        v1, v2 = get_float(val['V1']), get_float(val['V2'])
                        vc1, vc2 = get_float(val['Vc1']), get_float(val['Vc2'])
                        du_lieu_instance = DuLieuMotLanChay(v1, v2, vc1, vc2)
                    except ValueError:
                        du_lieu_instance = DuLieuMotLanChay(0.0, 0.0, 0.0, 0.0)

                    sai_so = round(_get_sai_so_dong_ho(du_lieu_instance), 2) if du_lieu_instance and _get_sai_so_dong_ho(du_lieu_instance) is not None else None
                    if hss is None:
                        hss = sai_so
                    else:
                        hss -= sai_so

                    sheet[f"AG{start_row}"] = sai_so if sai_so is not None else "-"
                    sheet[f"AG{start_row}"].alignment = center_align
                    sheet[f"AG{start_row}"].font = font_12
                    sheet[f"AG{start_row}"].number_format = 'General'

                    start_row += 1

                sheet[f"AJ{tmp_start_row}"] = round(hss, 2) if hss is not None else round(hieu_sai_so[index]['hss'], 2) 
                sheet[f"AJ{tmp_start_row}"].alignment = center_align
                sheet[f"AJ{tmp_start_row}"].font = font_12
                sheet[f"AJ{tmp_start_row}"].number_format = 'General'

                            
                # Tạo viền chấm cho toàn bộ vùng B2:AL3
                for row in sheet.iter_rows(min_row=tmp_start_row, max_row=tmp_start_row + len(lan_chay) - 1, min_col=2, max_col=38):
                    for cell in row:
                        cell.border = Border(
                            top=black_dotted,
                            bottom=black_dotted,
                            left=black_dotted,
                            right=black_dotted
                        )

                # Tạo viền liền đen cho các vùng bọc ngoài
                # Bọc ngoài B2:F3
                for row in sheet.iter_rows(min_row=tmp_start_row, max_row=tmp_start_row + len(lan_chay) - 1, min_col=2, max_col=6):
                    for cell in row:
                        if cell.row == tmp_start_row:  # Top border
                            cell.border = Border(top=black_solid, left=cell.border.left, right=cell.border.right, bottom=cell.border.bottom)
                        if cell.row == tmp_start_row + len(lan_chay) - 1:  # Bottom border
                            cell.border = Border(bottom=black_solid, left=cell.border.left, right=cell.border.right, top=cell.border.top)
                        if cell.column == 2:  # Left border
                            cell.border = Border(left=black_solid, top=cell.border.top, right=cell.border.right, bottom=cell.border.bottom)
                        if cell.column == 6:  # Right border
                            cell.border = Border(right=black_solid, top=cell.border.top, bottom=cell.border.bottom, left=cell.border.left)

                # Bọc ngoài G2:S3
                for row in sheet.iter_rows(min_row=tmp_start_row, max_row=tmp_start_row + len(lan_chay) - 1, min_col=7, max_col=19):
                    for cell in row:
                        if cell.row == tmp_start_row:  # Top border
                            cell.border = Border(top=black_solid, left=cell.border.left, right=cell.border.right, bottom=cell.border.bottom)
                        if cell.row == tmp_start_row + len(lan_chay) - 1:  # Bottom border
                            cell.border = Border(bottom=black_solid, left=cell.border.left, right=cell.border.right, top=cell.border.top)
                        if cell.column == 7:  # Left border
                            cell.border = Border(left=black_solid, top=cell.border.top, right=cell.border.right, bottom=cell.border.bottom)
                        if cell.column == 19:  # Right border
                            cell.border = Border(right=black_solid, top=cell.border.top, bottom=cell.border.bottom, left=cell.border.left)

                # Bọc ngoài T2:AF3
                for row in sheet.iter_rows(min_row=tmp_start_row, max_row=tmp_start_row + len(lan_chay) - 1, min_col=20, max_col=32):
                    for cell in row:
                        if cell.row == tmp_start_row:  # Top border
                            cell.border = Border(top=black_solid, left=cell.border.left, right=cell.border.right, bottom=cell.border.bottom)
                        if cell.row == tmp_start_row + len(lan_chay) - 1:  # Bottom border
                            cell.border = Border(bottom=black_solid, left=cell.border.left, right=cell.border.right, top=cell.border.top)
                        if cell.column == 20:  # Left border
                            cell.border = Border(left=black_solid, top=cell.border.top, right=cell.border.right, bottom=cell.border.bottom)
                        if cell.column == 32:  # Right border
                            cell.border = Border(right=black_solid, top=cell.border.top, bottom=cell.border.bottom, left=cell.border.left)

                # Bọc ngoài AG2:AL3
                for row in sheet.iter_rows(min_row=tmp_start_row, max_row=tmp_start_row + len(lan_chay) - 1, min_col=33, max_col=38):
                    for cell in row:
                        if cell.row == tmp_start_row:  # Top border
                            cell.border = Border(top=black_solid, left=cell.border.left, right=cell.border.right, bottom=cell.border.bottom)
                        if cell.row == tmp_start_row + len(lan_chay) - 1:  # Bottom border
                            cell.border = Border(bottom=black_solid, left=cell.border.left, right=cell.border.right, top=cell.border.top)
                        if cell.column == 33:  # Left border
                            cell.border = Border(left=black_solid, top=cell.border.top, right=cell.border.right, bottom=cell.border.bottom)
                        if cell.column == 38:  # Right border
                            cell.border = Border(right=black_solid, top=cell.border.top, bottom=cell.border.bottom, left=cell.border.left)

            
            sheet[f"D{start_row}"] = "Người thực hiện"
            sheet.merge_cells(f"D{start_row}:N{start_row}") 
            sheet[f"B{start_row + 4}"] = str(dongho_dict['nguoi_thuc_hien']).upper() if dongho_dict['nguoi_thuc_hien'] else ""
            sheet.merge_cells(f"B{start_row + 4}:P{start_row + 4}") 

            sheet[f"X{start_row}"] = "Người soát lại"
            sheet.merge_cells(f"X{start_row}:AF{start_row}") 
            sheet[f"V{start_row + 4}"] = str(dongho_dict['nguoi_soat_lai']).upper() if dongho_dict['nguoi_soat_lai'] else ""
            sheet.merge_cells(f"V{start_row + 4}:AH{start_row + 4}")  

            sheet[f"D{start_row}"].alignment = center_align
            sheet[f"D{start_row}"].font = font_13
            sheet[f"X{start_row}"].alignment = center_align
            sheet[f"X{start_row}"].font = font_13
            sheet[f"B{start_row + 4}"].alignment = center_align
            sheet[f"B{start_row + 4}"].font = font_13
            sheet[f"V{start_row + 4}"].alignment = center_align
            sheet[f"V{start_row + 4}"].font = font_13

            # Tạo một đối tượng BytesIO để lưu workbook
            excel_buffer = BytesIO()
            workbook.save(excel_buffer)
            excel_buffer.seek(0)

            workbook.close()
            return excel_buffer, fileName
    except Exception as e:
        print("Error creating sample BB Excel file: ", str(e))
        traceback.print_exc()
        return None, None      

@download_bp.route("/kiemdinh/bienban/<string:id>", methods=["GET"])
def get_excel_bb_kiem_dinh(id):
    try:
        decoded_id = decode(id)
        dongho = DongHo.query.filter_by(id=decoded_id).first_or_404()
        excel_buffer, fileName = _create_sample_bb_excel_buffer_file(dongho)

        if not excel_buffer:
            return jsonify({"msg": "Có lỗi xảy ra khi tạo file!"}), 500

        # Trả về file từ bộ nhớ
        return send_file(
            excel_buffer, 
            as_attachment=True, 
            download_name=fileName, 
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        # return send_file(f"../{bb_file_path}", as_attachment=True, download_name=fileName, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except NotFound:
        return jsonify({"msg": "Id không hợp lệ!"}), 404
    except Exception as e:
        return jsonify({"msg": f"Đã có lỗi xảy ra: {str(e)}"}), 500

@download_bp.route("/kiemdinh/bienban/pdf/<string:id>", methods=["GET"])
def get_pdf_bb_kiem_dinh(id):
    try:
        decoded_id = decode(id)
        dongho = DongHo.query.filter_by(id=decoded_id).first_or_404()
        excel_buffer, fileName = _create_sample_bb_excel_buffer_file(dongho)

        if not excel_buffer:
            return jsonify({"msg": "Có lỗi xảy ra khi tạo file!"}), 500

        pdf_buffer, pdf_filename = process_excel_bytesio_to_pdf(excel_buffer, fileName)
        
        if not pdf_buffer:
            return jsonify({"msg": "Có lỗi xảy ra khi tạo file!"}), 500

        # Trả về file từ bộ nhớ
        return send_file(
            pdf_buffer, 
            as_attachment=True, 
            download_name=fileName, 
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        # return send_file(f"../{bb_file_path}", as_attachment=True, download_name=fileName, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except NotFound:
        return jsonify({"msg": "Id không hợp lệ!"}), 404
    except Exception as e:
        print("BB: " + str(e))
        return jsonify({"msg": f"Đã có lỗi xảy ra: {str(e)}"}), 500

def _parse_date(date_str):
    try:
        return datetime.strptime(date_str, "%Y-%m-%d").date() if date_str else None
    except:
        return None

@download_bp.route("/kiemdinh/bienban/preview/pdf", methods=["POST"])
def preview_pdf_bb_kiem_dinh():
    try:
        data = request.get_json()

        if not data:
            return jsonify({"msg": "Thiếu thông tin đồng hồ!"}), 400

        # OPTIONAL: Nếu cần decode id:
        # decoded_id = decode(dongho_data["id"])

        # Tùy vào cấu trúc `dongho_data`, có thể tạo instance tạm:
        dongho = DongHo(
            data.get("ten_phuong_tien_do"),
            data.get("is_hieu_chuan"),
            data.get("index"),
            data.get("group_id"),

            data.get("sensor"),
            data.get("transitor"),
            data.get("serial"),

            data.get("co_so_san_xuat"),
            _parse_date(data.get("nam_san_xuat")),

            data.get("dn"),
            data.get("d"),
            data.get("ccx"),
            data.get("q3"),
            data.get("r"),
            data.get("qn"),
            data.get("k_factor"),
            data.get("so_qd_pdm"),

            data.get("so_tem"),
            data.get("so_giay_chung_nhan"),

            data.get("co_so_su_dung"),
            data.get("phuong_phap_thuc_hien"),
            data.get("chuan_thiet_bi_su_dung"),
            data.get("nguoi_thuc_hien"),
            _parse_date(data.get("ngay_thuc_hien")),

            data.get("dia_diem_thuc_hien"),

            data.get("ket_qua_check_vo_ngoai"),
            data.get("ket_qua_check_do_on_dinh_chi_so"),
            data.get("ket_qua_check_do_kin"),

            data.get("du_lieu_kiem_dinh"),
            data.get("nguoi_soat_lai"),

            _parse_date(data.get("hieu_luc_bien_ban")),
            data.get("last_updated"),

            data.get("noi_su_dung"),
            data.get("ten_khach_hang"),
            data.get("created_by"),
            data.get("nhiet_do"),
            data.get("do_am"),
            data.get("ma_quan_ly"),
        )


        excel_buffer, fileName = _create_sample_bb_excel_buffer_file(dongho, True)
        if not excel_buffer:
            return jsonify({"msg": "Có lỗi xảy ra khi tạo file!"}), 500

        pdf_buffer, pdf_filename = process_excel_bytesio_to_pdf(excel_buffer, fileName)
        if not pdf_buffer:
            return jsonify({"msg": "Có lỗi xảy ra khi tạo file!"}), 500

        return send_file(
            pdf_buffer,
            as_attachment=False,
            download_name=pdf_filename,
            mimetype="application/pdf"
        )

    except NotFound:
        return jsonify({"msg": "Không tìm thấy đồng hồ!"}), 404
    except Exception as e:
        print("Lỗi preview BB:", str(e))
        traceback.print_exc()
        return jsonify({"msg": f"Đã có lỗi xảy ra: {str(e)}"}), 500


def _create_sample_gcn_excel_buffer_file(dongho: DongHo, is_preview: bool = False):
    row_heights = 0.69 #cm
    try:
        dongho_dict = dongho.to_dict()
        if "du_lieu_kiem_dinh" in dongho_dict:
            try:
                dongho_dict["du_lieu_kiem_dinh"] = json.loads(
                    dongho_dict["du_lieu_kiem_dinh"]
                )
            except json.JSONDecodeError as e:
                print("Error decoding JSON: ", str(e))
                return None, None
        if dongho_dict['du_lieu_kiem_dinh']['ket_qua'] is None:
            print(">>> Không có dữ liệu kiểm định!")
            return None, None

        fileName = (
            "KĐ_GCN"
            + ("_" + dongho.so_giay_chung_nhan if dongho.so_giay_chung_nhan else "")
            + ("_" + dongho.ten_khach_hang if dongho.ten_khach_hang else "")
            # + ("_" + dongho.ten_dong_ho if dongho.ten_dong_ho else "")
            + ("_" + dongho.dn if dongho.dn else "")
            + ("_" + dongho.ngay_thuc_hien.strftime("%d-%m-%Y") if dongho.ngay_thuc_hien else "")
            + ".xlsx"
        )

        gcn_file_path = f"files/export/GCN/{fileName}"
        gcn_directory = os.path.dirname(gcn_file_path)

        # Ensure the directory exists
        if not os.path.exists(gcn_directory):
            os.makedirs(gcn_directory)

        if not os.path.exists(gcn_file_path):
            suffix = "Preview" if is_preview else "ExcelForm"
            src_file = f"files/GCN_{suffix}.xlsm"
            workbook = openpyxl.load_workbook(src_file)
            sheet = workbook.active

            sheet[f"Q10"] = f"FMS.KĐ.{dongho.so_giay_chung_nhan}.{dongho.ngay_thuc_hien.strftime('%y')}" if dongho.so_giay_chung_nhan and dongho.ngay_thuc_hien else ""
            sheet[f"J12"] = dongho.ten_phuong_tien_do if dongho.ten_phuong_tien_do else ""
            sheet[f"H14"] = dongho.co_so_san_xuat if dongho.co_so_san_xuat else ""
            sheet[f"H16"] = dongho.sensor if dongho.sensor else ""
            # sheet[f"H17"] = f"Chỉ thị: {dongho.transitor}" if dongho.transitor else ""    # co the không có
            # sheet[f"X16"] = dongho.seri_sensor if dongho.seri_sensor else ""      #khong co
            # sheet[f"X17"] = f"Chỉ thị: {dongho.seri_chi_thi}" if dongho.seri_chi_thi else ""      #khong co
            sheet[f"AA18"] = dongho.dn if dongho.dn else ""
            sheet[f"Y19"] = "Q3=" if dongho.q3 else "Qn="
            sheet[f"Z19"] = dongho.q3 if dongho.q3 else dongho.qn
            sheet[f"N20"] = "-" if dongho.ccx else ""
            sheet[f"N21"] = "-" if dongho.so_qd_pdm else ""
            sheet[f"N22"] = "-" if dongho.k_factor else ""
            sheet[f"O20"] = f"Cấp chính xác: {dongho.ccx}" if dongho.ccx else ""
            sheet[f"W20"] = f"Tỷ số Q3/Q1: (R) = {dongho.r}" if dongho.r else ""
            sheet[f"O21"] = "Ký hiệu PDM / Số quyết định:" if dongho.so_qd_pdm else ""
            sheet[f"AA21"] = dongho.so_qd_pdm if dongho.so_qd_pdm else ""
            sheet[f"O22"] = "Hệ số K: " if dongho.k_factor else ""
            sheet[f"S22"] = dongho.k_factor if dongho.k_factor else ""
            sheet[f"H23"] = dongho.co_so_su_dung if dongho.co_so_su_dung else ""
            sheet[f"H25"] = dongho.noi_su_dung if dongho.noi_su_dung else ""
            # sheet[f"H27"] = dongho.vi_tri if dongho.vi_tri else ""           # khong co
            sheet[f"H27"] = "" 
            sheet[f"L28"] = dongho.phuong_phap_thuc_hien if dongho.phuong_phap_thuc_hien else ""
            try:
                sheet[f"L29"] = PP_THUC_HIEN[dongho.phuong_phap_thuc_hien] if dongho.phuong_phap_thuc_hien else ""    
            except Exception as err:
                print(err)
                pass               
            sheet[f"F32"] = dongho.so_tem if dongho.so_tem else ""
            sheet[f"Z32"] = dongho.hieu_luc_bien_ban.strftime("%d/%m/%Y") if dongho.hieu_luc_bien_ban else ""
            sheet[f"S35"] = f"Hà Nội, ngày {dongho.ngay_thuc_hien.strftime('%d')} tháng {dongho.ngay_thuc_hien.strftime('%m')} năm {dongho.ngay_thuc_hien.strftime('%Y')}" if dongho.ngay_thuc_hien else ""
            sheet[f"A43"] = str(dongho.nguoi_thuc_hien).upper() if dongho.nguoi_thuc_hien else ""


            # Tạo một đối tượng BytesIO để lưu workbook
            excel_buffer = BytesIO()
            workbook.save(excel_buffer)
            excel_buffer.seek(0)

            workbook.close()
            return excel_buffer, fileName
    
    except Exception as e:
        print("Error creating sample BB Excel file: ", str(e))
        traceback.print_exc()
        return None, None
  
@download_bp.route("/kiemdinh/gcn/<string:id>", methods=["GET"])
def get_excel_gcn_kiem_dinh(id):
    try:
        decoded_id = decode(id)
        dongho = DongHo.query.filter_by(id=decoded_id).first_or_404()
        excel_buffer, fileName = _create_sample_gcn_excel_buffer_file(dongho)

        if not excel_buffer:
            return jsonify({"msg": "Có lỗi xảy ra khi tạo file!"}), 500

        # Trả về file từ bộ nhớ
        return send_file(
            excel_buffer, 
            as_attachment=True, 
            download_name=fileName, 
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        # return send_file(f"../{bb_file_path}", as_attachment=True, download_name=fileName, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except NotFound:
        return jsonify({"msg": "Id không hợp lệ!"}), 404
    except Exception as e:
        return jsonify({"msg": f"Đã có lỗi xảy ra: {str(e)}"}), 500

@download_bp.route("/kiemdinh/gcn/pdf/<string:id>", methods=["GET"])
def get_pdf_gcn_kiem_dinh(id):
    try:
        decoded_id = decode(id)
        dongho = DongHo.query.filter_by(id=decoded_id).first_or_404()
        excel_buffer, fileName = _create_sample_gcn_excel_buffer_file(dongho)

        if not excel_buffer:
            return jsonify({"msg": "Có lỗi xảy ra khi tạo file!"}), 500

        pdf_buffer, pdf_filename = process_excel_bytesio_to_pdf(excel_buffer, fileName)
        
        if not pdf_buffer:
            return jsonify({"msg": "Có lỗi xảy ra khi tạo file!"}), 500

        # Trả về file từ bộ nhớ
        return send_file(
            pdf_buffer, 
            as_attachment=True, 
            download_name=fileName, 
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        # return send_file(f"../{bb_file_path}", as_attachment=True, download_name=fileName, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except NotFound:
        return jsonify({"msg": "Id không hợp lệ!"}), 404
    except Exception as e:
        print("BB: " + str(e))
        return jsonify({"msg": f"Đã có lỗi xảy ra: {str(e)}"}), 500

@download_bp.route("/kiemdinh/gcn/preview/pdf", methods=["POST"])
def preview_pdf_gcn_kiem_dinh():
    try:
        data = request.get_json()

        if not data:
            return jsonify({"msg": "Thiếu thông tin đồng hồ!"}), 400

        # OPTIONAL: Nếu cần decode id:
        # decoded_id = decode(dongho_data["id"])

        # Tùy vào cấu trúc `dongho_data`, có thể tạo instance tạm:
        dongho = DongHo(
            data.get("ten_phuong_tien_do"),
            data.get("is_hieu_chuan"),
            data.get("index"),
            data.get("group_id"),

            data.get("sensor"),
            data.get("transitor"),
            data.get("serial"),

            data.get("co_so_san_xuat"),
            _parse_date(data.get("nam_san_xuat")),

            data.get("dn"),
            data.get("d"),
            data.get("ccx"),
            data.get("q3"),
            data.get("r"),
            data.get("qn"),
            data.get("k_factor"),
            data.get("so_qd_pdm"),

            data.get("so_tem"),
            data.get("so_giay_chung_nhan"),

            data.get("co_so_su_dung"),
            data.get("phuong_phap_thuc_hien"),
            data.get("chuan_thiet_bi_su_dung"),
            data.get("nguoi_thuc_hien"),
            _parse_date(data.get("ngay_thuc_hien")),

            data.get("dia_diem_thuc_hien"),

            data.get("ket_qua_check_vo_ngoai"),
            data.get("ket_qua_check_do_on_dinh_chi_so"),
            data.get("ket_qua_check_do_kin"),

            data.get("du_lieu_kiem_dinh"),
            data.get("nguoi_soat_lai"),

            _parse_date(data.get("hieu_luc_bien_ban")),
            data.get("last_updated"),

            data.get("noi_su_dung"),
            data.get("ten_khach_hang"),
            data.get("created_by"),
            data.get("nhiet_do"),
            data.get("do_am"),
            data.get("ma_quan_ly"),
        )


        excel_buffer, fileName = _create_sample_gcn_excel_buffer_file(dongho, True)
        if not excel_buffer:
            return jsonify({"msg": "Có lỗi xảy ra khi tạo file!"}), 500

        pdf_buffer, pdf_filename = process_excel_bytesio_to_pdf(excel_buffer, fileName)
        if not pdf_buffer:
            return jsonify({"msg": "Có lỗi xảy ra khi tạo file!"}), 500

        return send_file(
            pdf_buffer,
            as_attachment=False,
            download_name=pdf_filename,
            mimetype="application/pdf"
        )

    except NotFound:
        return jsonify({"msg": "Không tìm thấy đồng hồ!"}), 404
    except Exception as e:
        print("Lỗi preview BB:", str(e))
        traceback.print_exc()
        return jsonify({"msg": f"Đã có lỗi xảy ra: {str(e)}"}), 500








def _create_sample_hc_excel_buffer_file(dongho: DongHo, is_preview: bool = False):
    row_heights = 0.69 #cm
    try:
        dongho_dict = dongho.to_dict()
        if "du_lieu_kiem_dinh" in dongho_dict:
            try:
                dongho_dict["du_lieu_kiem_dinh"] = json.loads(
                    dongho_dict["du_lieu_kiem_dinh"]
                )
            except json.JSONDecodeError as e:
                return jsonify({"msg": "Có lỗi xảy ra khi trích xuất dữ liệu! Hãy thử lại"}), 400
        
        fileName = (
            "HC"
            + ("_" + dongho.so_giay_chung_nhan if dongho.so_giay_chung_nhan else "")
            + ("_" + dongho.ten_khach_hang if dongho.ten_khach_hang else "")
            # + ("_" + dongho.ten_dong_ho if dongho.ten_dong_ho else "")
            + ("_" + dongho.dn if dongho.dn else "")
            + ("_" + dongho.ngay_thuc_hien.strftime("%d-%m-%Y") if dongho.ngay_thuc_hien else "")
            + ".xlsx"
        )

        hc_file_path = f"files/export/HC/{fileName}"
        hc_directory = os.path.dirname(hc_file_path)

        # Ensure the directory exists
        if not os.path.exists(hc_directory):
            os.makedirs(hc_directory)

        if not os.path.exists(hc_file_path):
            suffix = "Preview" if is_preview else "ExcelForm"
            src_file = f"files/HC_{suffix}.xlsm"
            workbook = openpyxl.load_workbook(src_file)
            sheet = workbook.active
            so_giay = f"FMS.HC.{dongho.so_giay_chung_nhan}.{dongho.ngay_thuc_hien.strftime('%y')}" if dongho.so_giay_chung_nhan and dongho.ngay_thuc_hien else ""
            sheet[f"BL4"] = so_giay
            sheet[f"BB6"] = round(float(dongho.nhiet_do), 1) if isinstance(dongho.nhiet_do, (int, float)) else dongho.nhiet_do if dongho.nhiet_do else ""
            sheet[f"BJ6"] = round(float(dongho.do_am), 1) if isinstance(dongho.do_am, (int, float)) else dongho.do_am if dongho.do_am else ""

            sheet[f"Q8"] = so_giay
            sheet[f"J10"] = dongho.ten_phuong_tien_do if dongho.ten_phuong_tien_do else ""

            sheet[f"H12"] = dongho.co_so_san_xuat if dongho.co_so_san_xuat else ""
            sheet[f"Z12"] = "Mã Quản lý:" if dongho.ma_quan_ly else ""
            sheet[f"AA12"] = dongho.ma_quan_ly if dongho.ma_quan_ly else ""
            
            # sheet[f"H14"] = (f"Sensor: {dongho.sensor}" if dongho.transitor else dongho.sensor) if dongho.sensor else ""
            # sheet[f"H15"] = f"Chỉ thị: {dongho.transitor}" if dongho.transitor else ""

            # sheet[f"AA14"] = (f"Sensor: {dongho.seri_sensor}" if dongho.seri_chi_thi else dongho.seri_sensor) if dongho.seri_sensor else ""
            # sheet[f"AA15"] = f"Chỉ thị: {dongho.seri_chi_thi}" if dongho.seri_chi_thi else ""

            sheet[f"AA16"] = dongho.dn if dongho.dn else ""

            sheet[f"Y17"] = "Q3=" if dongho.q3 else "Qn="
            sheet[f"Z17"] = dongho.q3 if dongho.q3 else dongho.qn
            
            sheet[f"N18"] = "-" if dongho.ccx else ""
            sheet[f"O18"] = f"Cấp chính xác: {dongho.ccx}" if dongho.ccx else ""
            sheet[f"X18"] = f"Tỷ số Q3/Q1: (R) = {dongho.r}" if dongho.r else ""

            sheet[f"N19"] = "-" if dongho.k_factor else ""
            sheet[f"O19"] = f"Hệ số K: {dongho.k_factor}" if dongho.k_factor else ""

            sheet[f"H20"] = dongho.co_so_su_dung if dongho.co_so_su_dung else ""
            # TODO: check vitri
            # sheet[f"H21"] = dongho.vi_tri if dongho.vi_tri else ""           
            sheet[f"H22"] = "" 

            sheet[f"L23"] = dongho.phuong_phap_thuc_hien if dongho.phuong_phap_thuc_hien else ""
            sheet[f"L24"] = PP_THUC_HIEN[dongho.phuong_phap_thuc_hien] if dongho.phuong_phap_thuc_hien else ""   

            sheet[f"K25"] = dongho.chuan_thiet_bi_su_dung if dongho.chuan_thiet_bi_su_dung else ""   

            sheet[f"F27"] = dongho.so_tem if dongho.so_tem else ""

            sheet[f"AB27"] = dongho.hieu_luc_bien_ban.strftime("%d/%m/%Y") if dongho.hieu_luc_bien_ban else ""

            sheet[f"R30"] = f"Hà Nội, ngày {dongho.ngay_thuc_hien.strftime('%d')} tháng {dongho.ngay_thuc_hien.strftime('%m')} năm {dongho.ngay_thuc_hien.strftime('%Y')}" if dongho.ngay_thuc_hien else ""
            sheet[f"B39"] = str(dongho.nguoi_thuc_hien).upper() if dongho.nguoi_thuc_hien else ""

            
            du_lieu = dongho_dict['du_lieu_kiem_dinh']['du_lieu']
            hieu_sai_so = list(dongho_dict['du_lieu_kiem_dinh']['hieu_sai_so'])
            mf = list(dongho_dict['du_lieu_kiem_dinh']['mf'])
            titles = ["Q3","Q2","Q1"] if dongho.q3 else ['Qn', "Qt", "Qmin"] 

            sheet[f"AV11"] = du_lieu[titles[2]]['value'] if du_lieu[titles[2]]['value'] else "-"
            sheet[f"BA11"] = hieu_sai_so[2]['hss'] if hieu_sai_so[2]['hss'] != None or hieu_sai_so[2]['hss'] == 0  else "-" 
            sheet[f"BF11"] = mf[2]['mf'] if mf[2]['mf'] else "-" 

            sheet[f"AV13"] = du_lieu[titles[1]]['value'] if du_lieu[titles[1]]['value'] else "-"
            sheet[f"BA13"] = hieu_sai_so[1]['hss'] if hieu_sai_so[1]['hss'] != None or hieu_sai_so[1]['hss'] == 0  else "-" 
            sheet[f"BF13"] = mf[1]['mf'] if mf[1]['mf'] else "-" 

            sheet[f"AV15"] = du_lieu[titles[0]]['value'] if du_lieu[titles[0]]['value'] else "-"
            sheet[f"BA15"] = hieu_sai_so[0]['hss'] if hieu_sai_so[0]['hss'] != None or hieu_sai_so[0]['hss'] == 0  else "-" 
            sheet[f"BF15"] = mf[0]['mf'] if mf[0]['mf'] else "-" 

            desired_height = row_heights * 28.35
            # Run from row: 2
            for row in range(2, sheet.max_row + 1):
                sheet.row_dimensions[row].height = desired_height

            # Tạo một đối tượng BytesIO để lưu workbook
            excel_buffer = BytesIO()
            workbook.save(excel_buffer)
            excel_buffer.seek(0)

            workbook.close()
            return excel_buffer, fileName
    
    except Exception as e:
        print("Error creating sample BB Excel file: ", str(e))
        traceback.print_exc()
        return None, None
  
@download_bp.route("/hieuchuan/<string:id>", methods=["GET"])
def get_excel_hc_kiem_dinh(id):
    try:
        decoded_id = decode(id)
        dongho = DongHo.query.filter_by(id=decoded_id).first_or_404()
        excel_buffer, fileName = _create_sample_hc_excel_buffer_file(dongho)

        if not excel_buffer:
            return jsonify({"msg": "Có lỗi xảy ra khi tạo file!"}), 500

        # Trả về file từ bộ nhớ
        return send_file(
            excel_buffer, 
            as_attachment=True, 
            download_name=fileName, 
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        # return send_file(f"../{bb_file_path}", as_attachment=True, download_name=fileName, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except NotFound:
        return jsonify({"msg": "Id không hợp lệ!"}), 404
    except Exception as e:
        return jsonify({"msg": f"Đã có lỗi xảy ra: {str(e)}"}), 500

@download_bp.route("/hieuchuan/pdf/<string:id>", methods=["GET"])
def get_pdf_hc_kiem_dinh(id):
    try:
        decoded_id = decode(id)
        dongho = DongHo.query.filter_by(id=decoded_id).first_or_404()
        excel_buffer, fileName = _create_sample_hc_excel_buffer_file(dongho)

        if not excel_buffer:
            return jsonify({"msg": "Có lỗi xảy ra khi tạo file!"}), 500

        pdf_buffer, pdf_filename = process_excel_bytesio_to_pdf(excel_buffer, fileName)
        
        if not pdf_buffer:
            return jsonify({"msg": "Có lỗi xảy ra khi tạo file!"}), 500

        # Trả về file từ bộ nhớ
        return send_file(
            pdf_buffer, 
            as_attachment=True, 
            download_name=fileName, 
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        # return send_file(f"../{bb_file_path}", as_attachment=True, download_name=fileName, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except NotFound:
        return jsonify({"msg": "Id không hợp lệ!"}), 404
    except Exception as e:
        print("BB: " + str(e))
        return jsonify({"msg": f"Đã có lỗi xảy ra: {str(e)}"}), 500

@download_bp.route("/hieuchuan/preview/pdf", methods=["POST"])
def preview_pdf_hc_kiem_dinh():
    try:
        data = request.get_json()

        if not data:
            return jsonify({"msg": "Thiếu thông tin đồng hồ!"}), 400

        # OPTIONAL: Nếu cần decode id:
        # decoded_id = decode(dongho_data["id"])

        # Tùy vào cấu trúc `dongho_data`, có thể tạo instance tạm:
        dongho = DongHo(
            data.get("ten_phuong_tien_do"),
            data.get("is_hieu_chuan"),
            data.get("index"),
            data.get("group_id"),

            data.get("sensor"),
            data.get("transitor"),
            data.get("serial"),

            data.get("co_so_san_xuat"),
            _parse_date(data.get("nam_san_xuat")),

            data.get("dn"),
            data.get("d"),
            data.get("ccx"),
            data.get("q3"),
            data.get("r"),
            data.get("qn"),
            data.get("k_factor"),
            data.get("so_qd_pdm"),

            data.get("so_tem"),
            data.get("so_giay_chung_nhan"),

            data.get("co_so_su_dung"),
            data.get("phuong_phap_thuc_hien"),
            data.get("chuan_thiet_bi_su_dung"),
            data.get("nguoi_thuc_hien"),
            _parse_date(data.get("ngay_thuc_hien")),

            data.get("dia_diem_thuc_hien"),

            data.get("ket_qua_check_vo_ngoai"),
            data.get("ket_qua_check_do_on_dinh_chi_so"),
            data.get("ket_qua_check_do_kin"),

            data.get("du_lieu_kiem_dinh"),
            data.get("nguoi_soat_lai"),

            _parse_date(data.get("hieu_luc_bien_ban")),
            data.get("last_updated"),

            data.get("noi_su_dung"),
            data.get("ten_khach_hang"),
            data.get("created_by"),
            data.get("nhiet_do"),
            data.get("do_am"),
            data.get("ma_quan_ly"),
        )


        excel_buffer, fileName = _create_sample_hc_excel_buffer_file(dongho, True)
        if not excel_buffer:
            return jsonify({"msg": "Có lỗi xảy ra khi tạo file!"}), 500

        pdf_buffer, pdf_filename = process_excel_bytesio_to_pdf(excel_buffer, fileName)
        if not pdf_buffer:
            return jsonify({"msg": "Có lỗi xảy ra khi tạo file!"}), 500

        return send_file(
            pdf_buffer,
            as_attachment=False,
            download_name=pdf_filename,
            mimetype="application/pdf"
        )

    except NotFound:
        return jsonify({"msg": "Không tìm thấy đồng hồ!"}), 404
    except Exception as e:
        print("Lỗi preview BB:", str(e))
        traceback.print_exc()
        return jsonify({"msg": f"Đã có lỗi xảy ra: {str(e)}"}), 500









# @download_bp.route("/kiemdinh/hc/<string:id>", methods=["GET"])
# def get_hieu_chuan(id):
#     row_heights = 0.69 #cm

#     try:
#         decoded_id = decode(id)
#         dongho = DongHo.query.filter_by(id=decoded_id).first_or_404()
#         dongho_dict = dongho.to_dict()
#         if "du_lieu_kiem_dinh" in dongho_dict:
#             try:
#                 dongho_dict["du_lieu_kiem_dinh"] = json.loads(
#                     dongho_dict["du_lieu_kiem_dinh"]
#                 )
#             except json.JSONDecodeError as e:
#                 return jsonify({"msg": "Có lỗi xảy ra khi trích xuất dữ liệu! Hãy thử lại"}), 400
        
#         fileName = (
#             "HC"
#             + ("_" + dongho.so_giay_chung_nhan if dongho.so_giay_chung_nhan else "")
#             + ("_" + dongho.ten_khach_hang if dongho.ten_khach_hang else "")
#             # + ("_" + dongho.ten_dong_ho if dongho.ten_dong_ho else "")
#             + ("_" + dongho.dn if dongho.dn else "")
#             + ("_" + dongho.ngay_thuc_hien.strftime("%d-%m-%Y") if dongho.ngay_thuc_hien else "")
#             + ".xlsx"
#         )

#         hc_file_path = f"files/export/HC/{fileName}"
#         hc_directory = os.path.dirname(hc_file_path)

#         # Ensure the directory exists
#         if not os.path.exists(hc_directory):
#             os.makedirs(hc_directory)

#         if not os.path.exists(hc_file_path):
#             suffix = "Preview" if is_preview else "ExcelForm"
#             src_file = f"files/HC_{suffix}.xlsm"
#             workbook = openpyxl.load_workbook(src_file)
#             sheet = workbook.active
#             so_giay = f"FMS.HC.{dongho.so_giay_chung_nhan}.{dongho.ngay_thuc_hien.strftime('%y')}" if dongho.so_giay_chung_nhan and dongho.ngay_thuc_hien else ""
#             sheet[f"BL4"] = so_giay
#             sheet[f"BB6"] = round(float(dongho.nhiet_do), 1) if isinstance(dongho.nhiet_do, (int, float)) else dongho.nhiet_do if dongho.nhiet_do else ""
#             sheet[f"BJ6"] = round(float(dongho.do_am), 1) if isinstance(dongho.do_am, (int, float)) else dongho.do_am if dongho.do_am else ""

#             sheet[f"Q8"] = so_giay
#             sheet[f"J10"] = dongho.ten_phuong_tien_do if dongho.ten_phuong_tien_do else ""

#             sheet[f"H12"] = dongho.co_so_san_xuat if dongho.co_so_san_xuat else ""
#             sheet[f"Z12"] = "Mã Quản lý:" if dongho.ma_quan_ly else ""
#             sheet[f"AA12"] = dongho.ma_quan_ly if dongho.ma_quan_ly else ""
            
#             # sheet[f"H14"] = (f"Sensor: {dongho.sensor}" if dongho.transitor else dongho.sensor) if dongho.sensor else ""
#             # sheet[f"H15"] = f"Chỉ thị: {dongho.transitor}" if dongho.transitor else ""

#             # sheet[f"AA14"] = (f"Sensor: {dongho.seri_sensor}" if dongho.seri_chi_thi else dongho.seri_sensor) if dongho.seri_sensor else ""
#             # sheet[f"AA15"] = f"Chỉ thị: {dongho.seri_chi_thi}" if dongho.seri_chi_thi else ""

#             sheet[f"AA16"] = dongho.dn if dongho.dn else ""

#             sheet[f"Y17"] = "Q3=" if dongho.q3 else "Qn="
#             sheet[f"Z17"] = dongho.q3 if dongho.q3 else dongho.qn
            
#             sheet[f"N18"] = "-" if dongho.ccx else ""
#             sheet[f"O18"] = f"Cấp chính xác: {dongho.ccx}" if dongho.ccx else ""
#             sheet[f"X18"] = f"Tỷ số Q3/Q1: (R) = {dongho.r}" if dongho.r else ""

#             sheet[f"N19"] = "-" if dongho.k_factor else ""
#             sheet[f"O19"] = f"Hệ số K: {dongho.k_factor}" if dongho.k_factor else ""

#             sheet[f"H20"] = dongho.co_so_su_dung if dongho.co_so_su_dung else ""
#             sheet[f"H21"] = dongho.vi_tri if dongho.vi_tri else ""           
#             sheet[f"H22"] = "" 

#             sheet[f"L23"] = dongho.phuong_phap_thuc_hien if dongho.phuong_phap_thuc_hien else ""
#             sheet[f"L24"] = PP_THUC_HIEN[dongho.phuong_phap_thuc_hien] if dongho.phuong_phap_thuc_hien else ""   

#             sheet[f"K25"] = dongho.chuan_thiet_bi_su_dung if dongho.chuan_thiet_bi_su_dung else ""   

#             sheet[f"F27"] = dongho.so_tem if dongho.so_tem else ""

#             sheet[f"AB27"] = dongho.hieu_luc_bien_ban.strftime("%d/%m/%Y") if dongho.hieu_luc_bien_ban else ""

#             sheet[f"R30"] = f"Hà Nội, ngày {dongho.ngay_thuc_hien.strftime('%d')} tháng {dongho.ngay_thuc_hien.strftime('%m')} năm {dongho.ngay_thuc_hien.strftime('%Y')}" if dongho.ngay_thuc_hien else ""
#             sheet[f"B39"] = str(dongho.nguoi_thuc_hien).upper() if dongho.nguoi_thuc_hien else ""

            
#             du_lieu = dongho_dict['du_lieu_kiem_dinh']['du_lieu']
#             hieu_sai_so = list(dongho_dict['du_lieu_kiem_dinh']['hieu_sai_so'])
#             mf = list(dongho_dict['du_lieu_kiem_dinh']['mf'])
#             titles = ["Q3","Q2","Q1"] if dongho.q3 else ['Qn', "Qt", "Qmin"] 

#             sheet[f"AV11"] = du_lieu[titles[2]]['value'] if du_lieu[titles[2]]['value'] else "-"
#             sheet[f"BA11"] = hieu_sai_so[2]['hss'] if hieu_sai_so[2]['hss'] != None or hieu_sai_so[2]['hss'] == 0  else "-" 
#             sheet[f"BF11"] = mf[2]['mf'] if mf[2]['mf'] else "-" 

#             sheet[f"AV13"] = du_lieu[titles[1]]['value'] if du_lieu[titles[1]]['value'] else "-"
#             sheet[f"BA13"] = hieu_sai_so[1]['hss'] if hieu_sai_so[1]['hss'] != None or hieu_sai_so[1]['hss'] == 0  else "-" 
#             sheet[f"BF13"] = mf[1]['mf'] if mf[1]['mf'] else "-" 

#             sheet[f"AV15"] = du_lieu[titles[0]]['value'] if du_lieu[titles[0]]['value'] else "-"
#             sheet[f"BA15"] = hieu_sai_so[0]['hss'] if hieu_sai_so[0]['hss'] != None or hieu_sai_so[0]['hss'] == 0  else "-" 
#             sheet[f"BF15"] = mf[0]['mf'] if mf[0]['mf'] else "-" 

#             desired_height = row_heights * 28.35
#             # Run from row: 2
#             for row in range(2, sheet.max_row + 1):
#                 sheet.row_dimensions[row].height = desired_height

#             excel_buffer = BytesIO()
#             workbook.save(excel_buffer)
#             excel_buffer.seek(0)

#             workbook.close()
        
#         # Trả về file từ bộ nhớ thay vì lưu file vào hệ thống
#         return send_file(
#             excel_buffer, 
#             as_attachment=True, 
#             download_name=fileName, 
#             mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#         )
#         # return send_file(f"../{hc_file_path}", as_attachment=True, download_name=fileName, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
#     except NotFound:
#         return jsonify({"msg": "Id không hợp lệ!"}), 404
#     except Exception as e:
#         print("HC: " + str(e))
#         return jsonify({"msg": f"Đã có lỗi xảy ra: {str(e)}"}), 500

