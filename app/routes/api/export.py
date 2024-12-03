from flask import Blueprint, jsonify, send_file
from flask_jwt_extended import jwt_required
from app.models import DongHo
from app import db
from werkzeug.exceptions import NotFound
import json
from app.utils.url_encrypt import decode
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Border, Side
from app.constants import PP_THUC_HIEN

import os

export_bp = Blueprint("export", __name__)

class DuLieuMotLanChay:
    def __init__(self, V1: float, V2: float, Vc1: float, Vc2: float):
        self.V1 = V1
        self.V2 = V2
        self.Vc1 = Vc1
        self.Vc2 = Vc2

def get_sai_so_dong_ho(form_value: DuLieuMotLanChay) -> float:
    if form_value:
        # Kiểm tra và chuyển đổi các giá trị sang float
        try:
            V1 = float(form_value.V1) if form_value.V1 not in [None, ''] else 0.0
            V2 = float(form_value.V2) if form_value.V2 not in [None, ''] else 0.0
            Vc1 = float(form_value.Vc1) if form_value.Vc1 not in [None, ''] else 0.0
            Vc2 = float(form_value.Vc2) if form_value.Vc2 not in [None, ''] else 0.0
        except ValueError:
            return None  # Hoặc có thể trả về một giá trị mặc định khác

        if V2 == 0 and Vc2 == 0 and Vc1 == 0:
            return None

        VDHCT = V2 - V1
        VDHC = Vc2 - Vc1
        if VDHC != 0:
            error = ((VDHCT - VDHC) / VDHC) * 100
            return round(error, 3)  # Làm tròn đến 3 chữ số thập phân

    return None

@export_bp.route("/kiemdinh/bienban/<string:id>", methods=["GET"])
def get_bb_kiem_dinh(id):
    row_heights = 0.69 #cm

    try:
        decoded_id = decode(id)
        dongho = DongHo.query.filter_by(id=decoded_id).first_or_404()
        dongho_dict = dongho.to_dict()
        if "du_lieu_kiem_dinh" in dongho_dict:
            try:
                dongho_dict["du_lieu_kiem_dinh"] = json.loads(
                    dongho_dict["du_lieu_kiem_dinh"]
                )
            except json.JSONDecodeError as e:
                return jsonify({"msg": "Có lỗi xảy ra khi trích xuất dữ liệu! Hãy thử lại"}), 400
        if not dongho_dict['du_lieu_kiem_dinh']['ket_qua']:
            return jsonify({"msg": "Đồng hồ không đạt tiêu chuẩn xuất biên bản!"}), 422
        
        fileName = (
            "KĐ_BB"
            + ("_" + dongho.so_giay_chung_nhan if dongho.so_giay_chung_nhan else "")
            + ("_" + dongho.ten_khach_hang if dongho.ten_khach_hang else "")
            + ("_" + dongho.ten_dong_ho if dongho.ten_dong_ho else "")
            + ("_DN" + dongho.dn if dongho.dn else "")
            + ("_" + dongho.ngay_thuc_hien.strftime("%d-%m-%Y") if dongho.ngay_thuc_hien else "")
            + ".xlsx"
        )

        bb_file_path = f"excels/export/BB/{fileName}"
        if not os.path.exists(bb_file_path):
            src_file = "excels/BB_ExcelForm.xlsx"
            workbook = openpyxl.load_workbook(src_file)
            sheet = workbook.active

            positions = ["22", "23", "24", "26"]

            for pos in positions:
                check_img = Image("excels/data/check.png")
                sheet.add_image(check_img, f"U{pos}")
                check_img.width = 25
                check_img.height = 25

                circle_img = Image("excels/data/circle.png")
                sheet.add_image(circle_img, f"Z{pos}")
                circle_img.width = 25
                circle_img.height = 25

            sheet.cell(row=2, column=20, value="BIÊN BẢN KIỂM ĐỊNH")

            if dongho.so_giay_chung_nhan:
                sheet.cell(
                    row=3, column=20, value=f"FMS.KĐ.{dongho.so_giay_chung_nhan}.{dongho.ngay_thuc_hien.strftime('%y')}"
                )

            sheet.cell(
                row=5, column=8, value=dongho.phuong_tien_do if dongho.phuong_tien_do else ""
            )
            sheet.cell(
                row=6,
                column=8,
                value=dongho.co_so_san_xuat if dongho.co_so_san_xuat else "",
            )

            sheet.cell(
                row=7,
                column=8,
                value=(
                    (
                        f"Sensor: {dongho.kieu_sensor}"
                        if dongho.kieu_chi_thi
                        else dongho.kieu_sensor
                    )
                    if dongho.kieu_sensor
                    else ""
                ),
            )

            sheet.cell(
                row=8,
                column=8,
                value=(
                    f"Tranmistter: {dongho.kieu_chi_thi}" if dongho.kieu_chi_thi else ""
                ),
            )

            sheet.cell(
                row=7,
                column=27,
                value=(
                    (
                        f"Sensor: {dongho.seri_sensor}"
                        if dongho.seri_chi_thi
                        else dongho.seri_sensor
                    )
                    if dongho.seri_sensor
                    else ""
                ),
            )

            sheet.cell(
                row=8,
                column=27,
                value=(
                    f"Tranmistter: {dongho.seri_chi_thi}" if dongho.seri_chi_thi else ""
                ),
            )

            if dongho.dn:
                sheet.cell(row=9, column=26, value=dongho.dn)

            if dongho.q3 or dongho.qn:
                sheet.cell(row=10, column=24, value="Q3=" if dongho.q3 else "Qn=")
                sheet.cell(
                    row=10, column=25, value=dongho.q3 if dongho.q3 else dongho.qn
                )

            # M12
            if dongho.ccx:
                sheet.cell(row=11, column=13, value="-")
                sheet.cell(row=11, column=14, value=f"Cấp chính xác: {dongho.ccx}")

            if dongho.so_qd_pdm:
                sheet.cell(row=12, column=13, value="-")
                sheet.cell(row=12, column=14, value=f"Ký hiệu PDM / Số quyết định:")
                sheet.cell(row=12, column=26, value=dongho.so_qd_pdm)

            # R19
            if dongho.k_factor:
                sheet.cell(row=13, column=13, value="-")
                sheet.cell(row=13, column=14, value=f"Hệ số K:")
                sheet.cell(row=13, column=19, value=dongho.k_factor)
            # H7
            if dongho.co_so_su_dung:
                sheet.cell(row=14, column=8, value=dongho.noi_su_dung)

            # K10
            if dongho.phuong_phap_thuc_hien:
                sheet.cell(row=16, column=11, value=dongho.phuong_phap_thuc_hien)

            if dongho.chuan_thiet_bi_su_dung:
                sheet.cell(row=17, column=11, value=dongho.chuan_thiet_bi_su_dung)

            if dongho.nguoi_kiem_dinh:
                sheet.cell(row=19, column=9, value=str(dongho.nguoi_kiem_dinh).upper())

            sheet.cell(
                row=19,
                column=29,
                value=(
                    dongho.ngay_thuc_hien.strftime("%d/%m/%Y")
                    if dongho.ngay_thuc_hien
                    else ""
                ),
            )

            sheet[f"R31"] = "°C"
            sheet[f"AE31"] = "°C"

            sheet.cell(
                row=20,
                column=2,
                value=f"Địa điểm thực hiện: {dongho.noi_thuc_hien}" if dongho.noi_thuc_hien else "",
            )

            desired_height = row_heights * 28.35
            # Run from row: 2
            for row in range(2, sheet.max_row + 1):
                sheet.row_dimensions[row].height = desired_height

            # Bảng dl: start from row 32:
            du_lieu = dongho_dict['du_lieu_kiem_dinh']['du_lieu']
            hieu_sai_so = list(dongho_dict['du_lieu_kiem_dinh']['hieu_sai_so'])
            titles = ["Q3","Q2","Q1"] if dongho.q3 else ['Qn', "Qt", "Qmin"] 

            black_solid = Side(style="thin", color="000000")   # Viền liền màu đen (mỏng)
            black_dotted = Side(style="dotted", color="000000")

            start_row = 32
            for index, ll in enumerate(titles):
                tmp_start_row = start_row
                if ll == "Q1" or  ll == "Qmin":
                    ll_display = "I"
                elif ll == "Q2" or  ll == "Qt":
                    ll_display = "II"
                elif ll == "Q3" or  ll == "Qn":
                    ll_display = "III"
                else:
                    ll_display = ll

                dl_ll = du_lieu[ll]
                lan_chay = dict(dl_ll['lan_chay'])
                #merge cell
                sheet[f"B{start_row}"] = ll_display
                
                sheet[f"D{start_row}"] = (0.3 * (float(dl_ll['value']) if dl_ll['value'] else 0.0)) if ll == "Q3" else dl_ll['value']
                # sheet[f"AJ{start_row}"] = round(hieu_sai_so[index]['hss'], 1)
                sheet.merge_cells(f"B{start_row}:C{start_row + len(lan_chay) - 1}")    #title ll
                sheet.merge_cells(f"D{start_row}:F{start_row + len(lan_chay) - 1}")    #value ll
                sheet.merge_cells(f"AJ{start_row}:AL{start_row + len(lan_chay) - 1}")  #hss

                hss = None
                tmp_start_row = start_row
                for key, val in lan_chay.items():
                    sheet.merge_cells(f"G{start_row}:J{start_row}")
                    sheet[f"G{start_row}"] = val['V1']
                    sheet.merge_cells(f"K{start_row}:N{start_row}")
                    sheet[f"K{start_row}"] = val['V2']
                    sheet.merge_cells(f"O{start_row}:Q{start_row}")
                    try:
                        v1 = float(val['V1']) if val['V1'] else 0.0
                        v2 = float(val['V2']) if val['V2'] else 0.0
                        sheet[f"O{start_row}"] = v2 - v1 
                    except ValueError:
                        sheet[f"O{start_row}"] = "Lỗi"
                    sheet.merge_cells(f"R{start_row}:S{start_row}")
                    sheet[f"R{start_row}"] = val['Tdh']
                    sheet.merge_cells(f"T{start_row}:W{start_row}")
                    sheet[f"T{start_row}"] = val['Vc1']
                    sheet.merge_cells(f"X{start_row}:AA{start_row}")
                    sheet[f"X{start_row}"] = val['Vc2']
                    sheet.merge_cells(f"AB{start_row}:AD{start_row}")

                    try:
                        vc2 = float(val['Vc2']) if val['Vc2'] else 0.0
                        vc1 = float(val['Vc1']) if val['Vc1'] else 0.0
                        sheet[f"AB{start_row}"] = vc2 - vc1
                    except ValueError:
                        sheet[f"O{start_row}"] = "Lỗi"

                    sheet.merge_cells(f"AE{start_row}:AF{start_row}")
                    sheet[f"AE{start_row}"] = val['Tc']

                    sheet.merge_cells(f"AG{start_row}:AI{start_row}")
                    try:
                        v1 = float(val['V1']) if val['V1'] not in [None, ''] else 0.0
                        v2 = float(val['V2']) if val['V2'] not in [None, ''] else 0.0
                        vc1 = float(val['Vc1']) if val['Vc1'] not in [None, ''] else 0.0
                        vc2 = float(val['Vc2']) if val['Vc2'] not in [None, ''] else 0.0

                        du_lieu_instance = DuLieuMotLanChay(v1, v2, vc1, vc2)
                    except ValueError as err:
                        print(err)
                        du_lieu_instance = DuLieuMotLanChay(0.0, 0.0, 0.0, 0.0)
                        
                    if hss is None:
                        hss = round(get_sai_so_dong_ho(du_lieu_instance), 1)
                    else:
                        hss -= round(get_sai_so_dong_ho(du_lieu_instance), 1)
                    sai_so = round(get_sai_so_dong_ho(du_lieu_instance), 1)
                    sheet[f"AG{start_row}"] = sai_so if sai_so is not None else "Lỗi"
                    start_row += 1
                sheet[f"AJ{tmp_start_row}"] = hss or round(hieu_sai_so[index]['hss'], 1)
                            
                # Tạo viền chấm cho toàn bộ vùng B2:AL3
                for row in sheet.iter_rows(min_row=tmp_start_row, max_row=tmp_start_row + len(lan_chay) - 1, min_col=2, max_col=38):
                    for cell in row:
                        cell.border = Border(
                            top=black_dotted,
                            bottom=black_dotted,
                            left=black_dotted,
                            right=black_dotted
                        )

                # Tạo viền liền đen cho các vùng bọc ngoài
                # Bọc ngoài B2:F3
                for row in sheet.iter_rows(min_row=tmp_start_row, max_row=tmp_start_row + len(lan_chay) - 1, min_col=2, max_col=6):
                    for cell in row:
                        if cell.row == tmp_start_row:  # Top border
                            cell.border = Border(top=black_solid, left=cell.border.left, right=cell.border.right, bottom=cell.border.bottom)
                        if cell.row == tmp_start_row + len(lan_chay) - 1:  # Bottom border
                            cell.border = Border(bottom=black_solid, left=cell.border.left, right=cell.border.right, top=cell.border.top)
                        if cell.column == 2:  # Left border
                            cell.border = Border(left=black_solid, top=cell.border.top, right=cell.border.right, bottom=cell.border.bottom)
                        if cell.column == 6:  # Right border
                            cell.border = Border(right=black_solid, top=cell.border.top, bottom=cell.border.bottom, left=cell.border.left)

                # Bọc ngoài G2:S3
                for row in sheet.iter_rows(min_row=tmp_start_row, max_row=tmp_start_row + len(lan_chay) - 1, min_col=7, max_col=19):
                    for cell in row:
                        if cell.row == tmp_start_row:  # Top border
                            cell.border = Border(top=black_solid, left=cell.border.left, right=cell.border.right, bottom=cell.border.bottom)
                        if cell.row == tmp_start_row + len(lan_chay) - 1:  # Bottom border
                            cell.border = Border(bottom=black_solid, left=cell.border.left, right=cell.border.right, top=cell.border.top)
                        if cell.column == 7:  # Left border
                            cell.border = Border(left=black_solid, top=cell.border.top, right=cell.border.right, bottom=cell.border.bottom)
                        if cell.column == 19:  # Right border
                            cell.border = Border(right=black_solid, top=cell.border.top, bottom=cell.border.bottom, left=cell.border.left)

                # Bọc ngoài T2:AF3
                for row in sheet.iter_rows(min_row=tmp_start_row, max_row=tmp_start_row + len(lan_chay) - 1, min_col=20, max_col=32):
                    for cell in row:
                        if cell.row == tmp_start_row:  # Top border
                            cell.border = Border(top=black_solid, left=cell.border.left, right=cell.border.right, bottom=cell.border.bottom)
                        if cell.row == tmp_start_row + len(lan_chay) - 1:  # Bottom border
                            cell.border = Border(bottom=black_solid, left=cell.border.left, right=cell.border.right, top=cell.border.top)
                        if cell.column == 20:  # Left border
                            cell.border = Border(left=black_solid, top=cell.border.top, right=cell.border.right, bottom=cell.border.bottom)
                        if cell.column == 32:  # Right border
                            cell.border = Border(right=black_solid, top=cell.border.top, bottom=cell.border.bottom, left=cell.border.left)

                # Bọc ngoài AG2:AL3
                for row in sheet.iter_rows(min_row=tmp_start_row, max_row=tmp_start_row + len(lan_chay) - 1, min_col=33, max_col=38):
                    for cell in row:
                        if cell.row == tmp_start_row:  # Top border
                            cell.border = Border(top=black_solid, left=cell.border.left, right=cell.border.right, bottom=cell.border.bottom)
                        if cell.row == tmp_start_row + len(lan_chay) - 1:  # Bottom border
                            cell.border = Border(bottom=black_solid, left=cell.border.left, right=cell.border.right, top=cell.border.top)
                        if cell.column == 33:  # Left border
                            cell.border = Border(left=black_solid, top=cell.border.top, right=cell.border.right, bottom=cell.border.bottom)
                        if cell.column == 38:  # Right border
                            cell.border = Border(right=black_solid, top=cell.border.top, bottom=cell.border.bottom, left=cell.border.left)

            
            sheet[f"D{start_row}"] = "Người thực hiện"
            sheet.merge_cells(f"D{start_row}:N{start_row}") 
            sheet[f"E{start_row + 4}"] = str(dongho_dict['nguoi_kiem_dinh']).upper() if dongho_dict['nguoi_kiem_dinh'] else ""
            sheet.merge_cells(f"E{start_row + 4}:M{start_row + 4}") 

            sheet[f"X{start_row}"] = "Người soát lại"
            sheet.merge_cells(f"X{start_row}:AF{start_row}") 
            sheet[f"W{start_row + 4}"] = str(dongho_dict['nguoi_soat_lai']).upper() if dongho_dict['nguoi_soat_lai'] else ""
            sheet.merge_cells(f"W{start_row + 4}:AG{start_row + 4}")  

            # TODO: Save
            workbook.save(bb_file_path)
            workbook.close()
        return send_file(f"../{bb_file_path}", as_attachment=True, download_name=fileName, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except NotFound:
        return jsonify({"msg": "Id không hợp lệ!"}), 404
    except Exception as e:
        print("BB: " + str(e))
        return jsonify({"msg": f"Đã có lỗi xảy ra: {str(e)}"}), 500


@export_bp.route("/kiemdinh/gcn/<string:id>", methods=["GET"])
def get_gcn_kiem_dinh(id):
    row_heights = 0.69 #cm

    try:
        decoded_id = decode(id)
        dongho = DongHo.query.filter_by(id=decoded_id).first_or_404()
        dongho_dict = dongho.to_dict()
        if "du_lieu_kiem_dinh" in dongho_dict:
            try:
                dongho_dict["du_lieu_kiem_dinh"] = json.loads(
                    dongho_dict["du_lieu_kiem_dinh"]
                )
            except json.JSONDecodeError as e:
                return jsonify({"msg": "Có lỗi xảy ra khi trích xuất dữ liệu! Hãy thử lại"}), 400
        if not dongho_dict['du_lieu_kiem_dinh']['ket_qua']:
            return jsonify({"msg": "Đồng hồ không đạt tiêu chuẩn xuất biên bản!"}), 422
        
        fileName = (
            "KĐ_GCN"
            + ("_" + dongho.so_giay_chung_nhan if dongho.so_giay_chung_nhan else "")
            + ("_" + dongho.ten_khach_hang if dongho.ten_khach_hang else "")
            + ("_" + dongho.ten_dong_ho if dongho.ten_dong_ho else "")
            + ("_" + dongho.dn if dongho.dn else "")
            + ("_" + dongho.ngay_thuc_hien.strftime("%d-%m-%Y") if dongho.ngay_thuc_hien else "")
            + ".xlsx"
        )

        gcn_file_path = f"excels/export/GCN/{fileName}"
        if not os.path.exists(gcn_file_path):
            src_file = "excels/GCN_ExcelForm.xlsx"
            workbook = openpyxl.load_workbook(src_file)
            sheet = workbook.active

            sheet[f"Q10"] = f"FMS.KĐ.{dongho.so_giay_chung_nhan}.{dongho.ngay_thuc_hien.strftime('%y')}" if dongho.so_giay_chung_nhan and dongho.ngay_thuc_hien else ""
            sheet[f"J12"] = dongho.phuong_tien_do if dongho.phuong_tien_do else ""
            sheet[f"H14"] = dongho.co_so_san_xuat if dongho.co_so_san_xuat else ""
            sheet[f"H16"] = dongho.kieu_sensor if dongho.kieu_sensor else ""
            sheet[f"H17"] = f"Chỉ thị: {dongho.kieu_chi_thi}" if dongho.kieu_chi_thi else ""
            sheet[f"X16"] = dongho.seri_sensor if dongho.seri_sensor else ""
            sheet[f"X17"] = f"Chỉ thị: {dongho.seri_chi_thi}" if dongho.seri_chi_thi else ""
            sheet[f"AA18"] = dongho.dn if dongho.dn else ""
            sheet[f"Y19"] = "Q3=" if dongho.q3 else "Qn="
            sheet[f"Z19"] = dongho.q3 if dongho.q3 else dongho.qn
            sheet[f"N20"] = "-" if dongho.ccx else ""
            sheet[f"N21"] = "-" if dongho.so_qd_pdm else ""
            sheet[f"N22"] = "-" if dongho.k_factor else ""
            sheet[f"O20"] = f"Cấp chính xác: {dongho.ccx}" if dongho.ccx else ""
            sheet[f"W20"] = f"Tỷ số Q3/Q1: (R) = {dongho.r}" if dongho.r else ""
            sheet[f"O21"] = "Ký hiệu PDM / Số quyết định:" if dongho.so_qd_pdm else ""
            sheet[f"AA21"] = dongho.so_qd_pdm if dongho.so_qd_pdm else ""
            sheet[f"O22"] = "Hệ số K: " if dongho.k_factor else ""
            sheet[f"S22"] = dongho.k_factor if dongho.k_factor else ""
            sheet[f"H23"] = dongho.co_so_su_dung if dongho.co_so_su_dung else ""
            sheet[f"H25"] = dongho.noi_su_dung if dongho.noi_su_dung else ""
            sheet[f"H27"] = dongho.vi_tri if dongho.vi_tri else ""           
            sheet[f"H27"] = "" 
            sheet[f"L28"] = dongho.phuong_phap_thuc_hien if dongho.phuong_phap_thuc_hien else ""
            try:
                sheet[f"L29"] = PP_THUC_HIEN[dongho.phuong_phap_thuc_hien] if dongho.phuong_phap_thuc_hien else ""    
            except Exception as err:
                print(err)
                pass               
            sheet[f"F32"] = dongho.so_tem if dongho.so_tem else ""
            sheet[f"Z32"] = dongho.hieu_luc_bien_ban.strftime("%d/%m/%Y") if dongho.hieu_luc_bien_ban else ""
            sheet[f"S35"] = f"Hà Nội, ngày {dongho.ngay_thuc_hien.strftime('%d')} tháng {dongho.ngay_thuc_hien.strftime('%m')} năm {dongho.ngay_thuc_hien.strftime('%Y')}" if dongho.ngay_thuc_hien else ""
            sheet[f"A43"] = str(dongho.nguoi_kiem_dinh).upper() if dongho.nguoi_kiem_dinh else ""


            desired_height = row_heights * 28.35
            # Run from row: 2
            for row in range(2, sheet.max_row + 1):
                sheet.row_dimensions[row].height = desired_height

            # TODO: Save
            workbook.save(gcn_file_path)
            workbook.close()
        return send_file(f"../{gcn_file_path}", as_attachment=True, download_name=fileName, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except NotFound:
        return jsonify({"msg": "Id không hợp lệ!"}), 404
    except Exception as e:
        print("GCN: " + str(e))
        return jsonify({"msg": f"Đã có lỗi xảy ra: {str(e)}"}), 500

